from __future__ import annotations

from tabulate import tabulate
import copy
from typing import Tuple
from openpyxl import load_workbook, Workbook
class Table(object):
    def __init__(self):
        self.fields = []
        self.rows = []

        class Element(object):
            fields = []
            
            def __init__(self):
                self.lineno = -1

            def __hash__(self):
                return hash(str(self))
            def __len__(self):
                return len(Element.fields)
            
            def __getitem__(self, key):
                fname = Element.fields[key]
                return self.__dict__[fname]

            def __eq__(self, other):
                if len(self) == len(other):
                    
                    return all([s1 == s2 for (s1,s2) in zip(self, other)])
                else:
                    return False

            def __repr__(self):
                return tabulate([list(self)], headers = Element.fields, tablefmt = "grid")

        self.elt_type = Element

    def __len__(self): return len(self.rows)

    def __getitem__(self, key):
        return self.rows[key]

    def add_field(self, fname: str, fval: str | None = None):
        self.elt_type.fields.append(fname)
        self.fields.append(fname)
        setattr(self.elt_type, fname, fval)
        for r in self.rows:
            setattr(r, fname, fval)

    def remove_field(self, fname):
        delattr(self.elt_type, fname)
        self.fields.remove(fname)
        self.elt_type.fields.remove(fname)        
        for r in self.rows:
            delattr(r, fname)

    def new_elt(self, **kwargs):
        e = self.elt_type()
        for key, value in kwargs.items():
            e.__setattr__(key, value)
        return e
            
        return self.elt_type()

    def add_row(self, row: "Element"):
        self.rows.append(row)

    def copy(self):
        return copy.deepcopy(self)

    def filter(self, filter_fn):
        t = Table()
        t.fields = copy.deepcopy(self.fields)
        t.elt_type = copy.deepcopy(self.elt_type)
        for r in self.rows:
            if filter_fn(r):
                t.rows.append(copy.deepcopy(r))
        return t

    def uniquify(self: "Table"):
        s = set(self)
        self.rows = list(s)

    def check_shape(self, t2):
        if self.fields != t2.fields:
            raise ValueError("Tables must have same fields.")
            
    def union(self, t2):
        self.check_shape(t2)
        s1 = set(self)
        s2 = set(t2)
        su = s1.union(s2)
        t = Table()
        t.fields = copy.deepcopy(self.fields)
        t.elt_type = copy.deepcopy(self.fields)
        t.rows = copy.deepcopy(list(su))
        return t
        

    def __repr__(self):
        return tabulate(list([list(r) for r in self]), headers = self.fields, tablefmt = "grid")

    def from_xls(self, filename: str, sheet_name: str, offset: Tuple[int,int] = (0,0)):
        wb = load_workbook(filename, data_only = True)
        ws = wb[sheet_name]
        # field extraction
        for c in range(offset[1]+1,ws.max_column + 1):
            f = ws.cell(row = offset[0]+1, column = c).value
            self.add_field(f.strip())

        # elements extraction
        for r in range(offset[0]+2,ws.max_row + 1):
            e = self.new_elt()
            e.lineno = r
            for c in range(offset[1]+1,ws.max_column + 1):
                v = ws.cell(row = r, column = c).value
                e.__setattr__(self.fields[c -offset[1] -1], v)

            self.add_row(e)
                
            
            
            

        
